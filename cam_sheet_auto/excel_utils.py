import os
import sys  # 🔥 추가: sys 모듈 import
import openpyxl
import datetime
import traceback
from openpyxl.utils import column_index_from_string

# ✅ CAM SHEET.xlsx 파일 경로 설정
def get_template_path():
    """EXE 실행 시 CAM SHEET.xlsx 경로를 올바르게 반환"""
    if getattr(sys, 'frozen', False):  # EXE 실행 여부 확인
        base_path = sys._MEIPASS  # PyInstaller가 생성한 임시 폴더
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, "CAM SHEET.xlsx")

# ✅ CAM SHEET 템플릿 경로 설정
TEMPLATE_PATH = get_template_path()
print(f"📂 엑셀 템플릿 경로: {TEMPLATE_PATH}")

# ✅ 페이지별 데이터 입력 위치 (최대 24개씩)
PAGE_RANGES = [
    ("A6", "J29"),  # 1페이지 (6~29행, 좌측)
    ("A31", "J54"), # 2페이지 (31~54행, 좌측)
    ("K2", "T29"),  # 3페이지 (2~29행, 우측)
    ("K31", "T54")  # 4페이지 (31~54행, 우측)
]

def col_to_num(cell_address):
    """엑셀 열 문자(A, B, ... AA)를 숫자로 변환하는 함수"""
    col_str = ''.join(filter(str.isalpha, cell_address))  # A6 → A / AA6 → AA
    return column_index_from_string(col_str)  # A → 1, AA → 27

def convert_number(value):
    """📌 문자열 형태의 숫자를 실제 숫자로 변환"""
    try:
        if "." in value:  # 소수점 포함 → float 변환
            return float(value)
        return int(value)  # 정수 변환
    except ValueError:
        return value  # 변환 불가능한 경우 원래 값 유지

def set_value_in_merged_cell(sheet, row, col, value):
    """📌 병합된 셀인지 확인 후 첫 번째 셀에만 값 입력"""
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, _, _ = merged_range.bounds  # 병합된 첫 번째 셀 찾기
            first_cell = sheet.cell(row=min_row, column=min_col)
            first_cell.value = value  # ✅ 값 입력
            return
    cell.value = value  # ✅ 병합되지 않은 경우 값 입력

def get_unique_filename(folder_path, base_filename):
    """중복된 파일명이 있으면 -2, -3 식으로 카운팅하여 새로운 파일명 생성"""
    name, ext = os.path.splitext(base_filename)
    counter = 1
    new_filename = base_filename
    while os.path.exists(os.path.join(folder_path, new_filename)):
        counter += 1
        new_filename = f"{name}-{counter}{ext}"  # 파일명-2.xlsx, 파일명-3.xlsx 형식으로 변경
    return new_filename

def export_to_excel_with_auto_filename(job_number, machine_name, date, table_widget, folder_path):
    """PyQt UI 데이터를 받아서 CAM SHEET.xlsx에 저장 후 데이터 폴더에 자동 파일명으로 저장"""
    today_date = datetime.datetime.today().strftime("%m%d")  # ✅ MMDD 형식으로 변경
    base_filename = f"CAM_SHEET_{job_number}_{today_date}.xlsx"
    unique_filename = get_unique_filename(folder_path, base_filename)
    save_path = os.path.join(folder_path, unique_filename)

    

    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ 템플릿 파일을 찾을 수 없습니다! 현재 경로: {TEMPLATE_PATH}")
        return None

    try:
        workbook = openpyxl.load_workbook(TEMPLATE_PATH)
        sheet = workbook.active  # 첫 번째 시트 선택
        set_value_in_merged_cell(sheet, 3, 7, job_number)
        set_value_in_merged_cell(sheet, 3, 3, machine_name)
        set_value_in_merged_cell(sheet, 3, 10, date)

        total_rows = table_widget.rowCount()
        row_offset = 0
        for page_index, (start_cell, end_cell) in enumerate(PAGE_RANGES):
            if row_offset >= total_rows:
                break
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            start_col = col_to_num(start_cell)
            for i in range(24):
                if row_offset >= total_rows:
                    break
                file_name = table_widget.item(row_offset, 0).text() if table_widget.item(row_offset, 0) else ""
                tool_db = table_widget.item(row_offset, 1).text() if table_widget.item(row_offset, 1) else ""
                tool_number = table_widget.item(row_offset, 2).text() if table_widget.item(row_offset, 2) else ""
                allowance = table_widget.item(row_offset, 3).text() if table_widget.item(row_offset, 3) else ""
                work_content = table_widget.item(row_offset, 4).text() if table_widget.item(row_offset, 4) else ""
                sheet.cell(row=start_row + i, column=start_col, value=file_name)
                sheet.cell(row=start_row + i, column=start_col + 1, value=tool_db)
                sheet.cell(row=start_row + i, column=start_col + 4, value=convert_number(tool_number))
                sheet.cell(row=start_row + i, column=start_col + 5, value=convert_number(allowance))
                sheet.cell(row=start_row + i, column=start_col + 6, value=work_content)
                row_offset += 1
        workbook.save(save_path)
        workbook.close()
        print(f"✅ 엑셀 저장 완료: {save_path}")
        return save_path
    except Exception as e:
        error_message = f"❌ 파일 저장 중 오류 발생: {e}\n{traceback.format_exc()}"
        
        # CMD에서 강제 출력
        print(error_message)

        # 로그 파일 강제 생성
        try:
            with open("D:/error_log.txt", "w", encoding="utf-8") as f:
                f.write(error_message)
            print("✅ 오류 로그 저장 완료: D:/error_log.txt")
        except Exception as log_error:
            print(f"❌ 로그 파일 저장 실패: {log_error}")

        return None
